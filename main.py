from typing import List
from datetime import datetime
import os
import pandas as pd
import sys
from ubic import Ubic
from rel_ubic import ProdUbic
from openpyxl import load_workbook
import colors
from validation import validate_location, validate_roll, ERROR_MSG_LOCATION, ERROR_MSG_ROLL


def clear_screen():
    """Clear terminal screen."""
    os.system("clear")

def print_header():
    """Print application header."""
    clear_screen()
    print(f"{colors.CYAN}{'=' * 50}{colors.WHITE}")
    print(f"{colors.CYAN}     SISTEMA DE CONTROL DE ALMACÉN{colors.WHITE}")
    print(f"{colors.CYAN}{'=' * 50}{colors.WHITE}")
    print()

def print_separator():
    """Print visual separator."""
    print(f"{colors.BLUE}{'─' * 50}{colors.WHITE}")

def print_success(message: str):
    """Print success message."""
    print(f"{colors.GREEN}✓ {message}{colors.WHITE}")

def print_error(message: str):
    """Print error message."""
    print(f"{colors.RED}✗ {message}{colors.WHITE}")

def print_warning(message: str):
    """Print warning message."""
    print(f"{colors.YELLOW}⚠ {message}{colors.WHITE}")

def print_info(message: str):
    """Print info message."""
    print(f"{colors.CYAN}ℹ {message}{colors.WHITE}")

def load_existing_rolls(file_path: str) -> set:
    """Load all id_lote values from existing articulos sheet."""
    try:
        df = pd.read_excel(file_path, sheet_name='articulos', engine='openpyxl')
        if 'id_lote' in df.columns:
            return set(df['id_lote'].dropna().astype(str))
    except Exception:
        pass
    return set()

def main():
    # Generate filename with date and time: deposito_YYYY-MM-DD_HHMM.xlsx
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    file = f'deposito_{timestamp}.xlsx'
    user = None

    try:
        user = sys.argv[1]
    except Exception:
        if user is None:
            print_header()
            print_error("No ingresó un usuario válido.")
            print_info("Uso: python main.py <usuario>")
            return
    
    print_header()
    print_info(f"Usuario: {colors.WHITE}{user}")
    print_info(f"Archivo: {colors.WHITE}{file}")
    print_separator()

    try:
        book = load_workbook(file)
        print_success("Archivo de datos cargado correctamente.")
        existing_rolls = load_existing_rolls(file)
        if existing_rolls:
            print_info(f"Rollos existentes en sistema: {colors.WHITE}{len(existing_rolls)}")
        else:
            print_info("No hay rollos previos en el sistema.")
    except FileNotFoundError:
        print_warning(f"El archivo {file} no existe. Se creará uno nuevo.")
        df_ubic = pd.DataFrame(columns=['id', 'depo', 'pasillo', 'columna', 'nivel', 'ubic'])
        df_articulos = pd.DataFrame(columns=['id_usuario', 'id_ubicacion', 'id_lote', 'fecha_hora'])

        with pd.ExcelWriter(file, engine='openpyxl') as writer:
            df_ubic.to_excel(writer, sheet_name='ubicacion', index=False)
            df_articulos.to_excel(writer, sheet_name='articulos', index=False)
        existing_rolls = set()
    
    print()
    input("Presione ENTER para comenzar...")
    clear_screen()

    while True:
        print_header()
        print_info("UBICACIÓN")
        print_separator()
        ubicacion_input = input("📍 Escanee o ingrese la ubicación (10 caracteres)\n   Ingrese 0 para salir\n\n   INPUT: ")
        
        if ubicacion_input == "0":
            clear_screen()
            print_header()
            print_success(f"Hasta luego, {user}.")
            print_info("Sesión finalizada correctamente.")
            print()
            break
        
        if not validate_location(ubicacion_input):
            clear_screen()
            print_error(ERROR_MSG_LOCATION)
            print()
            input("Presione ENTER para continuar...")
            clear_screen()
            continue
        
        ubicacion = Ubic(ubicacion_input)
        df_ubicacion = pd.DataFrame([ubicacion.__dict__])
        
        with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            ubic_sheet = 'ubicacion'
            df_ubicacion.to_excel(
                writer,
                sheet_name=ubic_sheet,
                startrow=writer.sheets[ubic_sheet].max_row,
                index=False,
                header=False
            )
        
        print_success(f"Ubicación {ubicacion.id} registrada.")
        print_info(f"Depósito: {ubicacion.depo} | Pasillo: {ubicacion.pasillo} | Columna: {ubicacion.columna} | Nivel: {ubicacion.nivel}")
        print()
        input("Presione ENTER para comenzar a escanear artículos...")
        
        articulos:List[ProdUbic] = []
        clear_screen()
        
        while True:
            print_header()
            print_info("ARTÍCULOS")
            print_separator()
            print_info(f"UBICACIÓN: {colors.WHITE}{ubicacion.id}")
            print_info(f"ARTÍCULOS CARGADOS: {colors.WHITE}{len(articulos)}")
            print_separator()
            articulo_input = input("📦 Escanee o ingrese el código del artículo (7 caracteres)\n   Ingrese 0 para ver artículos cargados\n\n   INPUT: ")
            siguiente = False

            if articulo_input == "0":
                clear_screen()
                print_header()
                print_info("ARTÍCULOS CARGADOS EN ESTA UBICACIÓN")
                print_separator()
                
                if not articulos:
                    print_warning("No hay artículos cargados.")
                else:
                    for i, a in enumerate(articulos, 1):
                        print(f"{colors.CYAN}[{i}]{colors.WHITE}")
                        print(f"   USUARIO: {a.id_usuario}")
                        print(f"   UBICACIÓN: {a.id_ubicacion}")
                        print(f"   LOTE: {a.id_lote}")
                        print(f"   FECHA/HORA: {a.fecha_hora}")
                        print_separator()
                
                print()
                print_info(f"TOTAL: {colors.WHITE}{len(articulos)} artículos")
                print()
                respuesta = input("0. Confirmar y guardar\n1. Agregar más artículos\n\n   RESPUESTA: ")
                clear_screen()
                if respuesta == "0":
                    break
                elif respuesta == "1":
                    continue

            elif not validate_roll(articulo_input):
                clear_screen()
                print_error(ERROR_MSG_ROLL)
                print()
                input("Presione ENTER para continuar...")
                clear_screen()
                siguiente = True
            
            for a in articulos:
                if a.id_lote == articulo_input:
                    clear_screen()
                    print_error("Este artículo ya fue escaneado en esta sesión.")
                    print_info(f"LOTE: {colors.WHITE}{articulo_input}")
                    print()
                    input("Presione ENTER para continuar...")
                    clear_screen()
                    siguiente = True
            
            if articulo_input in existing_rolls:
                clear_screen()
                print_error("Este artículo ya existe en el sistema.")
                print_info(f"LOTE: {colors.WHITE}{articulo_input}")
                print_warning("No se permite re-ubicar artículos existentes.")
                print()
                input("Presione ENTER para continuar...")
                clear_screen()
                siguiente = True
            
            if siguiente:
                continue
            
            clear_screen()
            articulo = ProdUbic(user,ubicacion.id,articulo_input)
            
            articulos.append(articulo)
            existing_rolls.add(articulo_input)
            print_success(f"Artículo {articulo_input} registrado exitosamente.")
            print()
            input("Presione ENTER para continuar...")
            clear_screen()
        

        df_articulos = pd.DataFrame([a.__dict__ for a in articulos])

        print_header()
        print_info("RESUMEN DE ARTÍCULOS GUARDADOS")
        print_separator()
        print(df_articulos.to_string(index=False))
        print_separator()
        print()
        
        with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            art_sheet = 'articulos'
            df_articulos.to_excel(
                writer,
                sheet_name=art_sheet,
                startrow=writer.sheets[art_sheet].max_row,
                index=False,
                header=False
            )
        
        print_success(f"{len(articulos)} artículos guardados en {file}")
        print()
        input("Presione ENTER para continuar con la siguiente ubicación...")
        clear_screen()


main()
