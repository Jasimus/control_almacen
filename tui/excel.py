"""Direct openpyxl Excel I/O for warehouse control.

Replaces pandas-based operations with direct openpyxl read/write/append.
This reduces binary size by ~50MB and gives full control over the
workbook lifecycle.
"""

from typing import Optional
from openpyxl import Workbook, load_workbook
from ubic import Ubic
from rel_ubic import ProdUbic


class ExcelStore:
    """Manages Excel workbook operations for warehouse data.

    Handles loading or creating workbooks, appending location and article
    data, and querying existing rolls — all using openpyxl directly.
    """

    def __init__(self, filepath: str) -> None:
        """Initialize store and load or create the workbook.

        Args:
            filepath: Path to the .xlsx file.

        If the file does not exist, a new workbook is created with
        'ubicacion' and 'articulos' sheets with appropriate headers.
        """
        self.filepath = filepath
        try:
            self.wb = load_workbook(filepath)
        except FileNotFoundError:
            self.wb = Workbook()
            ubic_sheet = self.wb.active
            ubic_sheet.title = "ubicacion"
            ubic_sheet.append(["id", "depo", "pasillo", "columna", "nivel", "ubic"])

            art_sheet = self.wb.create_sheet("articulos")
            art_sheet.append(["id_usuario", "id_ubicacion", "id_lote", "fecha_hora"])

            self.wb.save(filepath)

    def append_ubicacion(self, ubic: Ubic) -> None:
        """Append a location row to the ubicacion sheet.

        Args:
            ubic: A Ubic instance with id, depo, pasillo, columna, nivel, ubic.
        """
        sheet = self.wb["ubicacion"]
        sheet.append([ubic.id, ubic.depo, ubic.pasillo, ubic.columna, ubic.nivel, ubic.ubic])
        self.wb.save(self.filepath)

    def append_articulos(self, articulos: list[ProdUbic]) -> None:
        """Append a batch of articles to the articulos sheet.

        Args:
            articulos: List of ProdUbic instances to append.
        """
        sheet = self.wb["articulos"]
        for a in articulos:
            sheet.append([a.id_usuario, a.id_ubic, a.id_lote, a.fecha_hora])
        self.wb.save(self.filepath)

    def load_existing_rolls(self) -> set[str]:
        """Load all id_lote values from existing articulos data.

        Returns:
            Set of all roll IDs (as strings) found in the articulos sheet.
        """
        sheet = self.wb["articulos"]
        rolls: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lote = row[2]  # id_lote is the third column
            if lote is not None:
                rolls.add(str(lote))
        return rolls
